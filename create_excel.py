import sys
import json
import openpyxl
from openpyxl.utils import get_column_letter

if len(sys.argv) < 3:
    print('Usage: python create_excel.py <json_data> <output_path>', file=sys.stderr)
    sys.exit(1)

try:
    data = json.loads(sys.argv[1])
except json.JSONDecodeError as e:
    print(f'JSONパースエラー: {e}', file=sys.stderr)
    sys.exit(1)

output_path = sys.argv[2]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '現金領収書'

# ヘッダー
ws.append(['日付', '店名', '金額', '消費税8%'])

# データ行
for item in data:
    ws.append([
        item.get('date', ''),
        item.get('store', ''),
        item['amount'] if item.get('amount') is not None else None,
        item['tax8'] if item.get('tax8') is not None else None,
    ])

# 合計行
data_len = len(data)
last_row = data_len + 1
sum_row = data_len + 2
ws.cell(row=sum_row, column=1, value='合計')
ws.cell(row=sum_row, column=3, value=f'=SUM(C2:C{last_row})')
ws.cell(row=sum_row, column=4, value=f'=SUM(D2:D{last_row})')

# 列幅
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14

wb.save(output_path)
print(f'saved: {output_path} | rows: {data_len}')
