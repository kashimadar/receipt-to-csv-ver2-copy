"""
参照用シート → Sheet2 データ転記スクリプト

【概要】
参照用シートの給与データ（所得税、支給額合計、非課税額合計、社会保険累計）を、
Sheet2の指定月区分の行に転記します。

【前提】
- openpyxl が必要: pip install openpyxl
- 参照用シートのフォーマット:
    行1: ヘッダー（社員名 例: "孫 成龍（00-0001）"）
    行2: 所得税
    行3: 課税額合計（※転記対象外）
    行4: 非課税額合計
    行5: 支給額合計
    行6: 社会保険累計
- Sheet2のフォーマット:
    行4: ヘッダー行
    行5以降: データ行（社員×月区分の組み合わせ）

【マッピング】
  参照用            →  Sheet2
  ─────────────────────────────────
  所得税(行2)       →  控除内訳_所得税 (P列)
  支給額合計(行5)   →  支給額合計 (F列)
  非課税額合計(行4)  →  支払内訳_内非課税分_金額_01 (G列)
  社会保険累計(行6)  →  控除内訳_社会保険料等_金額_01 (J列)
"""

import openpyxl
import unicodedata
import re
import sys
import os
from copy import copy


# ============================================================
# 設定
# ============================================================
WORK_DIR = r"C:\Users\sakur\Desktop\claude作業用"
INPUT_FILE = os.path.join(WORK_DIR, "明細一覧_成富商事株式会社_2026年2月分給与 PW;naritomi.xlsx")
OUTPUT_FILE = os.path.join(WORK_DIR, "output.xlsx")
TARGET_MONTH = "3月"
REF_SHEET_NAME = "参照用"
DATA_SHEET_NAME = "Sheet2"

# 参照用シートの行番号（1-indexed）
REF_ROW_HEADER = 1
REF_ROW_INCOME_TAX = 2
REF_ROW_NONTAX = 4
REF_ROW_TOTAL_PAY = 5
REF_ROW_SOCIAL_INS = 6

# Sheet2の列マッピング（参照用の行 → Sheet2の列名）
COLUMN_MAPPING = {
    REF_ROW_INCOME_TAX: "控除内訳_所得税",
    REF_ROW_TOTAL_PAY:  "支給額合計",
    REF_ROW_NONTAX:     "支払内訳_内非課税分_金額_01",
    REF_ROW_SOCIAL_INS: "控除内訳_社会保険料等_金額_01",
}


# ============================================================
# 名前正規化関数
# ============================================================
def normalize_name(name):
    if not name:
        return ""
    s = str(name)
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"※[^\s　]*", "", s)
    s = re.sub(r"[（(].+?[）)]", "", s)
    s = re.sub(r"[\s　]+", "", s)
    return s


def extract_english_name(name):
    if not name:
        return ""
    m = re.search(r"[（(]([A-Za-z\s　]+)[）)]", str(name))
    if m:
        return re.sub(r"[\s　]+", "", m.group(1)).upper()
    return ""


# ============================================================
# 類似文字マッピング（漢字の異体字対応）
# ============================================================
VARIANT_MAP = {
    "蓮": "連",
    "熙": "煕",
    "勳": "勲",
    "勛": "勲",
}


def apply_variants(name):
    result = name
    for old, new in VARIANT_MAP.items():
        result = result.replace(old, new)
    return result


# ============================================================
# メイン処理
# ============================================================
def main():
    print(f"ファイル読み込み中: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    if REF_SHEET_NAME not in wb.sheetnames:
        print(f"エラー: シート '{REF_SHEET_NAME}' が見つかりません。")
        sys.exit(1)
    if DATA_SHEET_NAME not in wb.sheetnames:
        print(f"エラー: シート '{DATA_SHEET_NAME}' が見つかりません。")
        sys.exit(1)

    ref_ws = wb[REF_SHEET_NAME]
    data_ws = wb[DATA_SHEET_NAME]

    # --- 参照用シートからデータ取得 ---
    print(f"\n=== 参照用シートからデータ読み込み ===")
    ref_data = {}
    ref_english = {}

    for col in range(2, ref_ws.max_column + 1):
        header = ref_ws.cell(row=REF_ROW_HEADER, column=col).value
        if not header:
            continue

        norm = normalize_name(header)
        norm_variant = apply_variants(norm)
        eng = extract_english_name(header)

        values = {}
        for ref_row, col_name in COLUMN_MAPPING.items():
            cell_val = ref_ws.cell(row=ref_row, column=col).value
            values[ref_row] = cell_val if cell_val is not None else 0

        ref_data[norm_variant] = values

        if eng:
            ref_english[eng] = norm_variant

    print(f"  参照用シートの社員数: {len(ref_data)}")

    # --- Sheet2のヘッダー解析 ---
    header_row = 4
    headers = {}
    for col in range(1, data_ws.max_column + 1):
        val = data_ws.cell(row=header_row, column=col).value
        if val:
            headers[val] = col

    col_name = headers.get("社員名")
    col_month = headers.get("月区分")
    col_targets = {}
    for ref_row, col_name_str in COLUMN_MAPPING.items():
        if col_name_str in headers:
            col_targets[ref_row] = headers[col_name_str]
        else:
            print(f"警告: Sheet2に列 '{col_name_str}' が見つかりません。")

    if not col_name or not col_month:
        print("エラー: Sheet2に '社員名' または '月区分' 列が見つかりません。")
        sys.exit(1)

    # --- マッチングと転記 ---
    print(f"\n=== 月区分 '{TARGET_MONTH}' の行にデータ転記 ===")
    matched_count = 0
    unmatched = []

    for row in range(header_row + 1, data_ws.max_row + 1):
        month_val = data_ws.cell(row=row, column=col_month).value
        if str(month_val).strip() != TARGET_MONTH:
            continue

        emp_name = data_ws.cell(row=row, column=col_name).value
        if not emp_name:
            continue

        norm = normalize_name(emp_name)
        norm_variant = apply_variants(norm)

        match_key = None

        if norm_variant in ref_data:
            match_key = norm_variant
        elif norm in ref_data:
            match_key = norm
        else:
            eng = extract_english_name(emp_name)
            if eng and eng in ref_english:
                match_key = ref_english[eng]
            else:
                eng_norm = re.sub(r"[\s　]+", "", str(emp_name)).upper()
                for ref_key in ref_data:
                    if ref_key in norm_variant or norm_variant in ref_key:
                        match_key = ref_key
                        break
                if not match_key:
                    for ref_eng, ref_jp in ref_english.items():
                        if ref_eng in eng_norm or eng_norm in ref_eng:
                            match_key = ref_jp
                            break

        if match_key:
            values = ref_data[match_key]
            for ref_row, target_col in col_targets.items():
                data_ws.cell(row=row, column=target_col).value = values[ref_row]
            matched_count += 1
            print(f"  ✓ 行{row}: {emp_name} → マッチ")
        else:
            unmatched.append((row, emp_name))

    # --- 結果サマリー ---
    print(f"\n=== 結果 ===")
    print(f"  マッチして転記: {matched_count} 名")
    print(f"  未マッチ: {len(unmatched)} 名")
    if unmatched:
        print(f"\n  【未マッチ社員一覧】")
        for row, name in unmatched:
            print(f"    行{row}: {name}")

    # --- 保存 ---
    print(f"\n保存中: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("完了！")


if __name__ == "__main__":
    main()