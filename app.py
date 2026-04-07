import streamlit as st
from pdf2image import convert_from_bytes
import pytesseract
from PIL import Image, ImageFilter, ImageEnhance
import re
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

st.set_page_config(page_title="現金領収書 → Excel", layout="wide")
st.title("現金領収書PDF → Excel 変換")
st.caption("スキャンしたPDFをアップロードすると、日付・店名・金額・消費税8%を読み取ります。")


# ── 画像前処理（OCR精度向上） ────────────────────────────────────
def preprocess(img: Image.Image) -> Image.Image:
    img = img.convert("L")
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)
    img = img.point(lambda x: 255 if x > 128 else 0, "1").convert("L")
    return img


# ── 和暦 → 西暦変換 ───────────────────────────────────────────────
def wareki_to_seireki(era: str, year: int) -> int:
    table = {"令和": 2018, "平成": 1988, "昭和": 1925, "R": 2018, "H": 1988, "S": 1925}
    base = table.get(era, 2018)
    return base + year


# ── 日付抽出 ──────────────────────────────────────────────────────
def extract_date(text: str) -> str:
    m = re.search(r"(20\d{2})[/\-年](\d{1,2})[/\-月](\d{1,2})", text)
    if m:
        return f"{m.group(1)}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"

    m = re.search(r"(令和|平成|昭和)[\s　]*(\d{1,2})年[\s　]*(\d{1,2})月[\s　]*(\d{1,2})日", text)
    if m:
        y = wareki_to_seireki(m.group(1), int(m.group(2)))
        return f"{y}/{int(m.group(3)):02d}/{int(m.group(4)):02d}"

    m = re.search(r"([RrHhSs令平昭])[\s　]*(\d{1,2})[./](\d{1,2})[./](\d{1,2})", text)
    if m:
        era_map = {"r": "令和", "R": "令和", "h": "平成", "H": "平成", "s": "昭和", "S": "昭和",
                   "令": "令和", "平": "平成", "昭": "昭和"}
        era = era_map.get(m.group(1), "令和")
        y = wareki_to_seireki(era, int(m.group(2)))
        return f"{y}/{int(m.group(3)):02d}/{int(m.group(4)):02d}"

    m = re.search(r"\b(\d{2})[./](\d{1,2})[./](\d{1,2})\b", text)
    if m:
        y = 2000 + int(m.group(1))
        return f"{y}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"

    return ""


# ── 金額抽出（合計金額） ──────────────────────────────────────────
def extract_amount(text: str):
    priority_patterns = [
        r"(?:合\s*計|お買上げ?金額|税込合計|ご請求金額|ご請求額|お支払[い]?金額)[\s　:：]*[¥￥]?\s*([\d,，]+)",
    ]
    for pat in priority_patterns:
        matches = re.findall(pat, text)
        if matches:
            vals = [int(v.replace(",", "").replace("，", "")) for v in matches]
            return max(vals)

    matches = re.findall(r"[¥￥]\s*([\d,，]+)", text)
    if matches:
        vals = [int(v.replace(",", "").replace("，", "")) for v in matches]
        return max(vals)

    matches = re.findall(r"([\d,，]{3,})\s*円", text)
    if matches:
        vals = [int(v.replace(",", "").replace("，", "")) for v in matches]
        return max(vals)

    return None


# ── 消費税8%抽出 ─────────────────────────────────────────────────
def extract_tax8(text: str):
    patterns = [
        r"(?:8[%％]対象消費税|消費税8[%％]|軽減税率.*?消費税|8[%％]税額)[\s　:：]*[¥￥]?\s*([\d,，]+)",
        r"(?:8[%％]消費税|税8[%％])[\s　:：]*[¥￥]?\s*([\d,，]+)",
        r"[※＊\*].*?消費税[\s　]*[¥￥]?\s*([\d,，]+)",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            return int(m.group(1).replace(",", "").replace("，", ""))
    return None


# ── 店名抽出 ──────────────────────────────────────────────────────
def extract_store(text: str) -> str:
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    skip = re.compile(r"^[\d\s　¥￥※＊\*\-\=\~\|]{3,}$|領収書|receipt|御領収|Tel|TEL|FAX|〒|郵便|ご利用", re.IGNORECASE)
    for line in lines[:10]:
        if len(line) >= 2 and not skip.search(line) and not re.match(r"^\d", line):
            return line
    return ""


# ── 1ページ分のパース ─────────────────────────────────────────────
def parse_receipt(text: str) -> dict:
    return {
        "日付": extract_date(text),
        "店名": extract_store(text),
        "金額": extract_amount(text),
        "消費税8%": extract_tax8(text),
    }


# ── Excel生成 ─────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "現金領収書"

    headers = ["日付", "店名", "金額", "消費税8%"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        ws.append([
            row.get("日付", ""),
            row.get("店名", ""),
            int(row["金額"]) if pd.notna(row.get("金額")) and row.get("金額") != "" else None,
            int(row["消費税8%"]) if pd.notna(row.get("消費税8%")) and row.get("消費税8%") != "" else None,
        ])

    n = len(df)
    ws.cell(row=n + 2, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=n + 2, column=3, value=f"=SUM(C2:C{n+1})")
    ws.cell(row=n + 2, column=4, value=f"=SUM(D2:D{n+1})")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── メインUI ─────────────────────────────────────────────────────
uploaded = st.file_uploader("PDFファイルをアップロード", type="pdf")

if uploaded:
    with st.spinner("OCR処理中...（ページ数によって数十秒かかります）"):
        try:
            images = convert_from_bytes(uploaded.read(), dpi=300)
        except Exception as e:
            st.error(f"PDF変換エラー: {e}")
            st.stop()

        records = []
        for i, img in enumerate(images):
            img = preprocess(img)
            text = pytesseract.image_to_string(
                img, lang="jpn",
                config="--psm 6 --oem 1"
            )
            record = parse_receipt(text)
            records.append(record)

    st.success(f"{len(records)} ページを読み取りました")

    df = pd.DataFrame(records)

    st.subheader("読み取り結果（直接編集できます）")
    edited = st.data_editor(
        df,
        num_rows="dynamic",
        column_config={
            "日付": st.column_config.TextColumn("日付", width=120),
            "店名": st.column_config.TextColumn("店名", width=300),
            "金額": st.column_config.NumberColumn("金額", format="%d"),
            "消費税8%": st.column_config.NumberColumn("消費税8%", format="%d"),
        },
        use_container_width=True,
    )

    filename = st.text_input("ファイル名", value="現金領収書一覧.xlsx")
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    excel_bytes = build_excel(edited)
    st.download_button(
        label="Excelダウンロード",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
